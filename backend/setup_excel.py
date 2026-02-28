import os
import random
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')
os.makedirs(DATA_DIR, exist_ok=True)
XLSX_PATH = os.path.join(DATA_DIR, 'device_inventory.xlsx')

# ---------------------------------------------------------------------------
# School / device config  (counts must match school_district.db)
# ---------------------------------------------------------------------------
SCHOOLS = [
    (1, 'Lincoln Elementary School',   68),
    (2, 'Jefferson Middle School',     95),
    (3, 'Washington High School',     100),
    (4, 'Roosevelt Elementary School', 57),
    (5, 'Kennedy Middle/High School',  85),
]

DEVICE_TYPES   = ['Chromebook', 'iPad', 'Windows Laptop', 'MacBook', 'Desktop PC', 'Smart Board']
DEVICE_WEIGHTS = [40, 20, 20, 5, 10, 5]

# Weighted toward active; 'retired' devices have no assigned_to
STATUSES = ['active', 'active', 'active', 'active', 'inactive', 'repair', 'retired']

# Realistic first + last names for assignment
FIRST_NAMES = [
    'James', 'Mary', 'Robert', 'Patricia', 'John', 'Jennifer', 'Michael',
    'Linda', 'William', 'Barbara', 'David', 'Susan', 'Richard', 'Jessica',
    'Joseph', 'Sarah', 'Thomas', 'Karen', 'Charles', 'Lisa', 'Christopher',
    'Nancy', 'Daniel', 'Betty', 'Matthew', 'Margaret', 'Anthony', 'Sandra',
    'Mark', 'Ashley', 'Donald', 'Dorothy', 'Steven', 'Kimberly', 'Paul',
    'Emily', 'Andrew', 'Donna', 'Joshua', 'Michelle',
]
LAST_NAMES = [
    'Smith', 'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Miller',
    'Davis', 'Rodriguez', 'Martinez', 'Hernandez', 'Lopez', 'Gonzalez',
    'Wilson', 'Anderson', 'Thomas', 'Taylor', 'Moore', 'Jackson', 'Martin',
    'Lee', 'Perez', 'Thompson', 'White', 'Harris', 'Sanchez', 'Clark',
    'Ramirez', 'Lewis', 'Robinson', 'Walker', 'Young', 'Allen', 'King',
    'Wright', 'Scott', 'Torres', 'Nguyen', 'Hill', 'Flores',
]

PREFIX_MAP = {
    'Chromebook':     'CB',
    'iPad':           'IP',
    'Windows Laptop': 'WL',
    'MacBook':        'MB',
    'Desktop PC':     'DC',
    'Smart Board':    'SB',
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def random_serial(school_id: int, dtype: str) -> str:
    chars = 'ABCDEFGHJKLMNPQRSTUVWXYZ0123456789'
    suffix = ''.join(random.choices(chars, k=8))
    return f"S{school_id:02d}-{PREFIX_MAP[dtype]}-{suffix}"


def random_last_active(status: str) -> str:
    if status == 'retired':
        days_ago = random.randint(180, 730)
    elif status == 'repair':
        days_ago = random.randint(7, 60)
    elif status == 'inactive':
        days_ago = random.randint(30, 180)
    else:
        days_ago = random.randint(0, 14)
    return (datetime.now() - timedelta(days=days_ago)).strftime('%Y-%m-%d')


def random_assignee(status: str) -> str:
    """Active/inactive devices are assigned to a person; repair = 'IT Dept'; retired = unassigned."""
    if status == 'retired':
        return 'Unassigned'
    if status == 'repair':
        return 'IT Department'
    return f"{random.choice(FIRST_NAMES)} {random.choice(LAST_NAMES)}"


# ---------------------------------------------------------------------------
# Build device rows
# ---------------------------------------------------------------------------

def build_inventory() -> list[dict]:
    rows = []
    for school_id, school_name, count in SCHOOLS:
        for _ in range(count):
            dtype  = random.choices(DEVICE_TYPES, weights=DEVICE_WEIGHTS, k=1)[0]
            status = random.choice(STATUSES)
            rows.append({
                'serial_number': random_serial(school_id, dtype),
                'school_name':   school_name,
                'device_type':   dtype,
                'status':        status,
                'last_active':   random_last_active(status),
                'assigned_to':   random_assignee(status),
            })
    return rows


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

HEADER_FILL   = PatternFill('solid', fgColor='1F4E79')   # dark blue
HEADER_FONT   = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
ALT_ROW_FILL  = PatternFill('solid', fgColor='D6E4F0')   # light blue
NORMAL_FILL   = PatternFill('solid', fgColor='FFFFFF')
SUMMARY_FILL  = PatternFill('solid', fgColor='E2EFDA')   # light green
SUMMARY_FONT  = Font(name='Calibri', bold=True, size=11)
BODY_FONT     = Font(name='Calibri', size=10)
CENTER        = Alignment(horizontal='center', vertical='center')
LEFT          = Alignment(horizontal='left',   vertical='center')

STATUS_COLORS = {
    'active':   'C6EFCE',   # green
    'inactive': 'FFEB9C',   # yellow
    'repair':   'FFCC99',   # orange
    'retired':  'FFC7CE',   # red
}

def thin_border() -> Border:
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)


def style_header_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = thin_border()


def style_data_cell(cell, alt: bool, align=LEFT):
    cell.fill      = ALT_ROW_FILL if alt else NORMAL_FILL
    cell.font      = BODY_FONT
    cell.alignment = align
    cell.border    = thin_border()


# ---------------------------------------------------------------------------
# Sheet 1 – Device Inventory
# ---------------------------------------------------------------------------

INV_HEADERS = ['serial_number', 'school_name', 'device_type', 'status', 'last_active', 'assigned_to']
COL_WIDTHS_INV = [22, 32, 18, 12, 14, 28]

def write_inventory_sheet(wb: openpyxl.Workbook, rows: list[dict]):
    ws = wb.active
    ws.title = 'Device Inventory'

    # Title row
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value     = 'School District – Device Inventory'
    title_cell.font      = Font(name='Calibri', bold=True, size=14, color='1F4E79')
    title_cell.alignment = CENTER
    title_cell.fill      = PatternFill('solid', fgColor='D6E4F0')
    ws.row_dimensions[1].height = 28

    # Timestamp row
    ws.merge_cells('A2:F2')
    ts_cell = ws['A2']
    ts_cell.value     = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ts_cell.font      = Font(name='Calibri', italic=True, size=9, color='595959')
    ts_cell.alignment = CENTER
    ws.row_dimensions[2].height = 16

    # Header
    for col, header in enumerate(INV_HEADERS, start=1):
        ws.cell(row=3, column=col, value=header.replace('_', ' ').title())
    style_header_row(ws, 3, len(INV_HEADERS))
    ws.row_dimensions[3].height = 20

    # Freeze header
    ws.freeze_panes = 'A4'

    # Data rows
    for i, row in enumerate(rows, start=1):
        excel_row = i + 3
        alt = (i % 2 == 0)
        values = [
            row['serial_number'],
            row['school_name'],
            row['device_type'],
            row['status'],
            row['last_active'],
            row['assigned_to'],
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            style_data_cell(cell, alt)

        # Colour-code the status cell
        status_cell = ws.cell(row=excel_row, column=4)
        color = STATUS_COLORS.get(row['status'], 'FFFFFF')
        status_cell.fill      = PatternFill('solid', fgColor=color)
        status_cell.alignment = CENTER

        ws.row_dimensions[excel_row].height = 15

    # Column widths
    for col, width in enumerate(COL_WIDTHS_INV, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Auto-filter on headers
    ws.auto_filter.ref = f"A3:F{3 + len(rows)}"


# ---------------------------------------------------------------------------
# Sheet 2 – Summary
# ---------------------------------------------------------------------------

SUMMARY_HEADERS = ['school_name', 'total_devices', 'active', 'inactive', 'under_repair']
COL_WIDTHS_SUM  = [32, 16, 12, 12, 16]

def build_summary(rows: list[dict]) -> list[dict]:
    summary = {}
    for school_id, school_name, _ in SCHOOLS:
        summary[school_name] = {'total': 0, 'active': 0, 'inactive': 0, 'repair': 0}

    for row in rows:
        s = summary[row['school_name']]
        s['total'] += 1
        if row['status'] == 'active':
            s['active'] += 1
        elif row['status'] == 'inactive':
            s['inactive'] += 1
        elif row['status'] == 'repair':
            s['repair'] += 1

    result = []
    for school_id, school_name, _ in SCHOOLS:
        d = summary[school_name]
        result.append({
            'school_name':  school_name,
            'total_devices': d['total'],
            'active':        d['active'],
            'inactive':      d['inactive'],
            'under_repair':  d['repair'],
        })
    return result


def write_summary_sheet(wb: openpyxl.Workbook, summary_rows: list[dict], total_devices: int):
    ws = wb.create_sheet('Summary')

    # Title
    ws.merge_cells('A1:E1')
    t = ws['A1']
    t.value     = 'Device Inventory – School Summary'
    t.font      = Font(name='Calibri', bold=True, size=14, color='375623')
    t.alignment = CENTER
    t.fill      = PatternFill('solid', fgColor='E2EFDA')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:E2')
    ts = ws['A2']
    ts.value     = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    ts.font      = Font(name='Calibri', italic=True, size=9, color='595959')
    ts.alignment = CENTER
    ws.row_dimensions[2].height = 16

    # Header
    header_labels = ['School Name', 'Total Devices', 'Active', 'Inactive', 'Under Repair']
    for col, label in enumerate(header_labels, start=1):
        ws.cell(row=3, column=col, value=label)
    style_header_row(ws, 3, len(SUMMARY_HEADERS))
    ws.row_dimensions[3].height = 20

    ws.freeze_panes = 'A4'

    # Data rows
    for i, row in enumerate(summary_rows, start=1):
        excel_row = i + 3
        alt = (i % 2 == 0)
        values = [
            row['school_name'],
            row['total_devices'],
            row['active'],
            row['inactive'],
            row['under_repair'],
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.fill      = SUMMARY_FILL if not alt else PatternFill('solid', fgColor='F0F7EC')
            cell.font      = BODY_FONT
            cell.alignment = CENTER if col > 1 else LEFT
            cell.border    = thin_border()

        ws.row_dimensions[excel_row].height = 18

    # Totals row
    totals_row = len(summary_rows) + 4
    ws.cell(row=totals_row, column=1, value='DISTRICT TOTAL')
    for col, key in enumerate(['total_devices', 'active', 'inactive', 'under_repair'], start=2):
        ws.cell(row=totals_row, column=col,
                value=sum(r[key] for r in summary_rows))
    for col in range(1, 6):
        cell = ws.cell(row=totals_row, column=col)
        cell.fill      = PatternFill('solid', fgColor='375623')
        cell.font      = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        cell.alignment = CENTER if col > 1 else LEFT
        cell.border    = thin_border()
    ws.row_dimensions[totals_row].height = 20

    # Column widths
    for col, width in enumerate(COL_WIDTHS_SUM, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def setup():
    random.seed(42)

    inventory = build_inventory()
    summary   = build_summary(inventory)

    wb = openpyxl.Workbook()
    write_inventory_sheet(wb, inventory)
    write_summary_sheet(wb, summary, len(inventory))

    wb.save(XLSX_PATH)

    print(f"Excel file created at: {os.path.abspath(XLSX_PATH)}\n")
    print(f"  Sheet 1 – Device Inventory: {len(inventory)} rows")
    print(f"  Sheet 2 – Summary:\n")
    print(f"  {'School':<35} {'Total':>5} {'Active':>6} {'Inactive':>8} {'Repair':>7}")
    print('  ' + '-' * 65)
    for row in summary:
        print(f"  {row['school_name']:<35} {row['total_devices']:>5} "
              f"{row['active']:>6} {row['inactive']:>8} {row['under_repair']:>7}")
    totals = {k: sum(r[k] for r in summary) for k in ['total_devices', 'active', 'inactive', 'under_repair']}
    print('  ' + '-' * 65)
    print(f"  {'TOTAL':<35} {totals['total_devices']:>5} "
          f"{totals['active']:>6} {totals['inactive']:>8} {totals['under_repair']:>7}")


if __name__ == '__main__':
    setup()
