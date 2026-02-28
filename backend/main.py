import asyncio
import json
import logging
import os
import re
import sqlite3
import uuid
from datetime import datetime, timedelta
from pathlib import Path

import anthropic
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
import openpyxl

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
load_dotenv()

BASE_DIR   = Path(__file__).parent
DATA_DIR   = BASE_DIR / '..' / 'data'
FRONTEND_DIR = BASE_DIR / '..' / 'frontend'
DB_PATH    = DATA_DIR / 'school_district.db'
XLSX_PATH  = DATA_DIR / 'device_inventory.xlsx'

client = anthropic.Anthropic(api_key=os.environ['ANTHROPIC_API_KEY'])

# ---------------------------------------------------------------------------
# Tool implementations
# ---------------------------------------------------------------------------

def parse_excel() -> list[dict]:
    """Read device_inventory.xlsx and return per-school summary statistics."""
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    # Use the Summary sheet if it exists, otherwise compute from Device Inventory
    if 'Summary' in wb.sheetnames:
        ws = wb['Summary']
        rows = list(ws.iter_rows(values_only=True))
        # Find header row
        header_row_idx = None
        for i, row in enumerate(rows):
            if row and row[0] and str(row[0]).strip().lower() in ('school name', 'school_name'):
                header_row_idx = i
                break

        if header_row_idx is not None:
            headers = [str(h).strip().lower().replace(' ', '_') if h else '' for h in rows[header_row_idx]]
            results = []
            for row in rows[header_row_idx + 1:]:
                if not row or not row[0]:
                    continue
                val = str(row[0]).strip()
                if val.upper() in ('DISTRICT TOTAL', 'TOTAL', ''):
                    continue
                record = dict(zip(headers, row))
                results.append({
                    'school_name':  val,
                    'total_devices': int(record.get('total_devices') or record.get('total') or 0),
                    'active':        int(record.get('active', 0) or 0),
                    'inactive':      int(record.get('inactive', 0) or 0),
                    'under_repair':  int(record.get('under_repair', 0) or 0),
                })
            wb.close()
            return results

    # Fallback: compute from Device Inventory sheet
    ws = wb['Device Inventory']
    rows = list(ws.iter_rows(values_only=True))
    header_row_idx = None
    for i, row in enumerate(rows):
        if row and 'serial' in str(row[0]).lower():
            header_row_idx = i
            break

    if header_row_idx is None:
        wb.close()
        return []

    headers = [str(h).strip().lower().replace(' ', '_') if h else '' for h in rows[header_row_idx]]
    school_col  = headers.index('school_name')
    status_col  = headers.index('status')

    tally: dict[str, dict] = {}
    for row in rows[header_row_idx + 1:]:
        if not row or not row[school_col]:
            continue
        school = str(row[school_col]).strip()
        status = str(row[status_col]).strip().lower() if row[status_col] else ''
        if school not in tally:
            tally[school] = {'total': 0, 'active': 0, 'inactive': 0, 'under_repair': 0}
        tally[school]['total'] += 1
        if status == 'active':
            tally[school]['active'] += 1
        elif status == 'inactive':
            tally[school]['inactive'] += 1
        elif status == 'repair':
            tally[school]['under_repair'] += 1

    wb.close()
    return [
        {'school_name': k, **v}
        for k, v in tally.items()
    ]


def query_database(query_type: str, school_name: str | None = None) -> list[dict]:
    """Query school_district.db and return results as a list of dicts."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    try:
        if query_type == 'device_summary':
            sql = '''
                SELECT s.name AS school_name,
                       COUNT(d.id)                        AS total_devices,
                       SUM(d.status = 'active')           AS active,
                       SUM(d.status = 'inactive')         AS inactive,
                       SUM(d.status = 'repair')           AS under_repair,
                       SUM(d.status = 'retired')          AS retired
                FROM schools s
                LEFT JOIN devices d ON d.school_id = s.id
            '''
            params: list = []
            if school_name:
                sql += ' WHERE s.name = ?'
                params.append(school_name)
            sql += ' GROUP BY s.id ORDER BY s.name'

        elif query_type == 'license_summary':
            sql = '''
                SELECT s.name AS school_name,
                       l.software_name,
                       l.total_seats,
                       l.used_seats,
                       l.expiry_date,
                       CAST(julianday(l.expiry_date) - julianday('now') AS INTEGER) AS days_until_expiry
                FROM licenses l
                JOIN schools s ON s.id = l.school_id
            '''
            params = []
            if school_name:
                sql += ' WHERE s.name = ?'
                params.append(school_name)
            sql += ' ORDER BY s.name, l.expiry_date'

        elif query_type == 'inactive_devices':
            sql = '''
                SELECT s.name AS school_name,
                       d.serial_number,
                       d.device_type,
                       d.status,
                       d.last_active
                FROM devices d
                JOIN schools s ON s.id = d.school_id
                WHERE d.status IN ('inactive', 'repair', 'retired')
            '''
            params = []
            if school_name:
                sql += ' AND s.name = ?'
                params.append(school_name)
            sql += ' ORDER BY s.name, d.status, d.last_active'

        elif query_type == 'expiring_licenses':
            cutoff = (datetime.now() + timedelta(days=90)).strftime('%Y-%m-%d')
            sql = '''
                SELECT s.name AS school_name,
                       l.software_name,
                       l.total_seats,
                       l.used_seats,
                       l.expiry_date,
                       CAST(julianday(l.expiry_date) - julianday('now') AS INTEGER) AS days_until_expiry
                FROM licenses l
                JOIN schools s ON s.id = l.school_id
                WHERE l.expiry_date <= ?
            '''
            params = [cutoff]
            if school_name:
                sql += ' AND s.name = ?'
                params.append(school_name)
            sql += ' ORDER BY l.expiry_date'

        else:
            raise ValueError(f"Unknown query_type: {query_type!r}")

        cur.execute(sql, params)
        return [dict(row) for row in cur.fetchall()]

    finally:
        conn.close()


def generate_report(school_name: str, analysis_text: str) -> dict:
    """Save a markdown report to the data folder and return its path + content."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_name = re.sub(r'[^a-zA-Z0-9]+', '_', school_name).strip('_')
    filename  = f"report_{safe_name}_{timestamp}.md"
    filepath  = DATA_DIR / filename

    header = f"""# IT Infrastructure Report – {school_name}

**Generated:** {datetime.now().strftime('%B %d, %Y at %I:%M %p')}
**District:** Northview Unified School District

---

"""
    full_content = header + analysis_text

    filepath.write_text(full_content, encoding='utf-8')

    return {
        'file_path': str(filepath.resolve()),
        'filename':  filename,
        'content':   full_content,
    }


# ---------------------------------------------------------------------------
# Email draft store  (in-memory; keyed by draft_id)
# ---------------------------------------------------------------------------

# Each entry: { 'to', 'subject', 'body', 'attachment_path', 'status', 'created_at' }
email_drafts: dict[str, dict] = {}


def draft_email(
    to: list[str],
    subject: str,
    body: str,
    attachment_path: str | None = None,
) -> dict:
    """Store an email draft and return its id + payload. Does NOT send."""
    draft_id = str(uuid.uuid4())
    draft = {
        'draft_id':        draft_id,
        'to':              to,
        'subject':         subject,
        'body':            body,
        'attachment_path': attachment_path,
        'status':          'pending_approval',
        'created_at':      datetime.now().isoformat(),
    }
    email_drafts[draft_id] = draft
    logging.info(f"[email] Draft created: {draft_id} → {to}")
    return draft


def send_email(draft_id: str) -> dict:
    """Mark a draft as sent (simulated — no real email is sent)."""
    draft = email_drafts.get(draft_id)
    if not draft:
        return {'error': f'Draft {draft_id!r} not found'}
    draft['status'] = 'sent'
    draft['sent_at'] = datetime.now().isoformat()
    logging.info(
        f"[email] SENT (simulated) draft={draft_id}  "
        f"to={draft['to']}  subject={draft['subject']!r}"
    )
    return {
        'success': True,
        'draft_id': draft_id,
        'to': draft['to'],
        'subject': draft['subject'],
        'sent_at': draft['sent_at'],
        'message': f"Email '{draft['subject']}' sent to {', '.join(draft['to'])} (simulated).",
    }


# ---------------------------------------------------------------------------
# Anthropic tool definitions
# ---------------------------------------------------------------------------

TOOLS = [
    {
        'name': 'parse_excel',
        'description': (
            'Read the device_inventory.xlsx Excel file and return per-school summary '
            'statistics including total device count, active, inactive, and under-repair counts. '
            'Use this when the user asks about device inventory data from Excel, '
            'wants an overview of all schools, or needs a quick device count summary.'
        ),
        'input_schema': {
            'type': 'object',
            'properties': {},
            'required': [],
        },
    },
    {
        'name': 'query_database',
        'description': (
            'Query the school_district SQLite database for detailed IT data. '
            'Supports four query types:\n'
            '- "device_summary": total, active, inactive, repair, and retired counts per school\n'
            '- "license_summary": all software licenses with seat counts and expiry dates\n'
            '- "inactive_devices": list of devices with inactive, repair, or retired status\n'
            '- "expiring_licenses": licenses expiring within the next 90 days\n'
            'Optionally filter by school_name to narrow results to one school.'
        ),
        'input_schema': {
            'type': 'object',
            'properties': {
                'query_type': {
                    'type': 'string',
                    'enum': ['device_summary', 'license_summary', 'inactive_devices', 'expiring_licenses'],
                    'description': 'The type of data to retrieve from the database.',
                },
                'school_name': {
                    'type': 'string',
                    'description': (
                        'Optional. Exact school name to filter results. '
                        'Valid values: "Lincoln Elementary School", "Jefferson Middle School", '
                        '"Washington High School", "Roosevelt Elementary School", '
                        '"Kennedy Middle/High School".'
                    ),
                },
            },
            'required': ['query_type'],
        },
    },
    {
        'name': 'generate_report',
        'description': (
            'Generate a formatted markdown report for a school, save it to the data folder, '
            'and return the file path and full content. '
            'Call this after gathering and analyzing data when the user asks for a report, '
            'summary document, or wants to export findings. '
            'The analysis_text should be complete markdown with sections, bullet points, '
            'and actionable recommendations.'
        ),
        'input_schema': {
            'type': 'object',
            'properties': {
                'school_name': {
                    'type': 'string',
                    'description': 'The name of the school the report is about, or "District" for a district-wide report.',
                },
                'analysis_text': {
                    'type': 'string',
                    'description': (
                        'Full markdown body of the report. Include sections like ## Device Health, '
                        '## License Status, ## Recommendations. Use real numbers from the data. '
                        'Be specific and actionable.'
                    ),
                },
            },
            'required': ['school_name', 'analysis_text'],
        },
    },
    {
        'name': 'draft_email',
        'description': (
            'Prepare an email draft with recipient(s), subject, body, and optional report attachment. '
            'This does NOT send the email — it stores the draft and returns a draft_id for human review. '
            'Use this when the user asks to email or send a report. '
            'After calling this tool, always tell the user to review and approve the email before it is sent. '
            'Never call send_email directly after draft_email — wait for human approval.'
        ),
        'input_schema': {
            'type': 'object',
            'properties': {
                'to': {
                    'type': 'array',
                    'items': {'type': 'string'},
                    'description': 'List of recipient email addresses.',
                },
                'subject': {
                    'type': 'string',
                    'description': 'Email subject line.',
                },
                'body': {
                    'type': 'string',
                    'description': 'Full email body text. Be professional and concise.',
                },
                'attachment_path': {
                    'type': 'string',
                    'description': 'Optional. Absolute file path to a report .md file to attach.',
                },
            },
            'required': ['to', 'subject', 'body'],
        },
    },
    {
        'name': 'send_email',
        'description': (
            'Send a previously drafted email identified by draft_id. '
            'IMPORTANT: Only call this after receiving explicit human approval via the /approve-email endpoint. '
            'Do not call this tool proactively or automatically after draft_email.'
        ),
        'input_schema': {
            'type': 'object',
            'properties': {
                'draft_id': {
                    'type': 'string',
                    'description': 'The draft_id returned by the draft_email tool.',
                },
            },
            'required': ['draft_id'],
        },
    },
]

SYSTEM_PROMPT = """You are an expert K-12 school district IT analyst assistant. \
You help IT administrators understand and manage their technology infrastructure.

CRITICAL RULES:
- ALWAYS use the available tools to retrieve real data before answering any question about devices, licenses, or schools.
- NEVER invent, estimate, or guess numbers. Every statistic you state must come from a tool call.
- If you need data from multiple sources, make multiple tool calls.
- Be specific and actionable — name exact schools, exact counts, exact dates.
- When asked for a report or summary document, use generate_report to save it.
- Format your final responses clearly with headers, bullet points, and tables where helpful.
- Flag risks proactively: devices sitting inactive too long, licenses expiring soon, utilization above 90%.

EMAIL WORKFLOW (follow this exactly when asked to send or email a report):
1. First generate the report using generate_report if not already done.
2. Use draft_email to prepare the email — include the report filename in the body and set attachment_path if available.
3. STOP and tell the user: "I've prepared an email draft for your review. Please approve or reject it above before I send it."
4. NEVER call send_email on your own. Only send_email is triggered by explicit human approval via the UI.
5. If the user asks you to send without prior approval, always draft first and ask for approval.
"""

# ---------------------------------------------------------------------------
# Tool dispatcher
# ---------------------------------------------------------------------------

def dispatch_tool(tool_name: str, tool_input: dict) -> str:
    try:
        if tool_name == 'parse_excel':
            result = parse_excel()
        elif tool_name == 'query_database':
            result = query_database(**tool_input)
        elif tool_name == 'generate_report':
            result = generate_report(**tool_input)
        elif tool_name == 'draft_email':
            result = draft_email(**tool_input)
        elif tool_name == 'send_email':
            result = send_email(**tool_input)
        else:
            result = {'error': f'Unknown tool: {tool_name}'}
        return json.dumps(result, default=str)
    except Exception as exc:
        return json.dumps({'error': str(exc)})


# ---------------------------------------------------------------------------
# FastAPI app
# ---------------------------------------------------------------------------

app = FastAPI(title='School District IT Assistant')

app.add_middleware(
    CORSMiddleware,
    allow_origins=['*'],
    allow_methods=['*'],
    allow_headers=['*'],
)


class ChatRequest(BaseModel):
    message: str


# ---------------------------------------------------------------------------
# SSE streaming helper
# ---------------------------------------------------------------------------

def sse(payload: dict) -> str:
    return f"data: {json.dumps(payload)}\n\n"


async def run_agent(message: str):
    """Agentic loop: yield SSE strings until Claude produces a final text response."""

    yield sse({'type': 'thinking', 'content': 'Analyzing your request...'})
    await asyncio.sleep(0)

    messages: list[dict] = [{'role': 'user', 'content': message}]

    while True:
        # Call Claude (synchronous SDK call run in thread pool to avoid blocking)
        response = await asyncio.to_thread(
            client.messages.create,
            model='claude-sonnet-4-6',
            max_tokens=4096,
            system=SYSTEM_PROMPT,
            tools=TOOLS,
            messages=messages,
        )

        # Collect text and tool-use blocks
        text_parts: list[str] = []
        tool_calls: list[anthropic.types.ToolUseBlock] = []

        for block in response.content:
            if block.type == 'text':
                text_parts.append(block.text)
            elif block.type == 'tool_use':
                tool_calls.append(block)

        # ── Tool calls ───────────────────────────────────────────────────────
        if tool_calls:
            # Append Claude's full response (may contain both text and tool_use blocks)
            messages.append({'role': 'assistant', 'content': response.content})

            tool_results = []
            for tc in tool_calls:
                # Friendly label for the UI
                label_map = {
                    'parse_excel':     'Reading Excel inventory file...',
                    'query_database':  f"Querying database ({tc.input.get('query_type', '')})...",
                    'generate_report': f"Generating report for {tc.input.get('school_name', '')}...",
                    'draft_email':     f"Drafting email to {', '.join(tc.input.get('to', []))}...",
                    'send_email':      'Sending email...',
                }
                yield sse({'type': 'tool_call', 'content': label_map.get(tc.name, f"Calling {tc.name}...")})
                await asyncio.sleep(0)

                raw_result = await asyncio.to_thread(dispatch_tool, tc.name, tc.input)

                # Parse result for UI feedback
                try:
                    parsed = json.loads(raw_result)
                except Exception:
                    parsed = {}

                # Emit special SSE events for email flow
                if tc.name == 'draft_email' and isinstance(parsed, dict) and 'draft_id' in parsed:
                    yield sse({'type': 'email_draft', 'content': parsed})
                    await asyncio.sleep(0)
                    snippet = f"Email draft prepared (id: {parsed['draft_id'][:8]}…)"
                elif tc.name == 'send_email' and isinstance(parsed, dict) and parsed.get('success'):
                    yield sse({'type': 'email_sent', 'content': parsed})
                    await asyncio.sleep(0)
                    snippet = f"Email sent to {', '.join(parsed.get('to', []))}"
                elif isinstance(parsed, list):
                    snippet = f"Retrieved {len(parsed)} record(s)."
                elif isinstance(parsed, dict) and 'filename' in parsed:
                    snippet = f"Report saved: {parsed['filename']}"
                elif isinstance(parsed, dict) and 'error' in parsed:
                    snippet = f"Error: {parsed['error']}"
                else:
                    snippet = 'Done.'

                yield sse({'type': 'tool_result', 'content': snippet})
                await asyncio.sleep(0)

                tool_results.append({
                    'type':        'tool_result',
                    'tool_use_id': tc.id,
                    'content':     raw_result,
                })

            messages.append({'role': 'user', 'content': tool_results})
            # Continue the loop so Claude can process results and respond
            continue

        # ── Final text response ──────────────────────────────────────────────
        final_text = '\n'.join(text_parts).strip()
        if not final_text:
            final_text = 'I was unable to produce a response. Please try again.'

        yield sse({'type': 'final', 'content': final_text})
        yield sse({'type': 'done'})
        return


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@app.get('/health')
async def health():
    return {'status': 'ok'}


@app.get('/')
async def serve_frontend():
    index = FRONTEND_DIR / 'index.html'
    if not index.exists():
        raise HTTPException(status_code=404, detail='Frontend not found')
    return FileResponse(str(index))


@app.get('/download/{filename}')
async def download_file(filename: str):
    # Prevent path traversal
    filepath = (DATA_DIR / filename).resolve()
    if not str(filepath).startswith(str(DATA_DIR.resolve())):
        raise HTTPException(status_code=400, detail='Invalid filename')
    if not filepath.exists():
        raise HTTPException(status_code=404, detail='File not found')
    return FileResponse(str(filepath), filename=filename)


@app.post('/approve-email/{draft_id}')
async def approve_email(draft_id: str):
    draft = email_drafts.get(draft_id)
    if not draft:
        raise HTTPException(status_code=404, detail=f'Draft {draft_id!r} not found')
    if draft['status'] == 'sent':
        return {'success': True, 'message': 'Email was already sent.', 'draft': draft}
    result = await asyncio.to_thread(send_email, draft_id)
    if 'error' in result:
        raise HTTPException(status_code=500, detail=result['error'])
    return result


@app.post('/chat')
async def chat(request: ChatRequest):
    if not request.message.strip():
        raise HTTPException(status_code=400, detail='Message cannot be empty')

    return StreamingResponse(
        run_agent(request.message),
        media_type='text/event-stream',
        headers={
            'Cache-Control':   'no-cache',
            'X-Accel-Buffering': 'no',
        },
    )


# ---------------------------------------------------------------------------
# Dev entry-point
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    import uvicorn
    uvicorn.run('main:app', host='0.0.0.0', port=8000, reload=True)
